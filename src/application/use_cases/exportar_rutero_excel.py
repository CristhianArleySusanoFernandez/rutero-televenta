from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.domain.ports.llamada_repository import LlamadaRepository
from src.domain.servicios.dia_visita import fecha_del_dia_en_semana

# Mismo orden y nombres que el Excel original — así el archivo se puede
# volver a subir sin que el parser tenga que adivinar nada.
COLUMNAS = [
    ("Usuario", "usuario_id"),
    ("ASESOR", "asesor_campo"),
    ("Cod Cliente", "cod_cliente"),
    ("Documento", "documento"),
    ("Cliente", "nombre"),
    ("Razon social", "razon_social"),
    ("Direccion", "direccion"),
    ("Barrio", "barrio"),
    ("Ciudad", "ciudad"),
    ("Dias Visita", "dias_visita"),
    ("Telefono", "telefono"),
    ("Novedades", "novedad_excel"),
]

COLUMNA_COD_CLIENTE = 3  # 1-indexado: debe salir SIEMPRE como texto


class ExportarRuteroExcel:
    """
    Exporta el rutero completo de la semana (lunes a sábado) de un asesor,
    con las 12 columnas del Excel original, para que se pueda corregir y
    volver a subir. Distinto de ExportarReporte (que exporta solo
    excepciones del día con 6 columnas) — no comparten datos ni lógica.
    """

    def __init__(self, llamada_repo: LlamadaRepository):
        self._llamada_repo = llamada_repo

    async def execute(self, fecha_ancla: date, asesor: str) -> tuple[BytesIO, str]:
        lunes = fecha_del_dia_en_semana(fecha_ancla, 0)
        sabado = fecha_del_dia_en_semana(fecha_ancla, 5)

        clientes = await self._llamada_repo.get_clientes_semana(lunes, sabado, asesor)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rutero"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, (encabezado, _) in enumerate(COLUMNAS, 1):
            cell = ws.cell(row=1, column=col, value=encabezado)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        usuario_id = ""
        for row_idx, cliente in enumerate(clientes, 2):
            if not usuario_id and cliente.get("usuario_id"):
                usuario_id = cliente["usuario_id"]
            for col, (_, campo) in enumerate(COLUMNAS, 1):
                valor = cliente.get(campo) or ""
                cell = ws.cell(row=row_idx, column=col, value=str(valor) if valor != "" else "")
                if col == COLUMNA_COD_CLIENTE:
                    # Crítico: cod_cliente trae ceros a la izquierda
                    # (ej. "000373587"). Forzar formato texto para que
                    # Excel nunca lo reinterprete como número al abrirlo.
                    cell.number_format = "@"

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        usuario_para_nombre = usuario_id or "SINUSUARIO"
        nombre_archivo = f"RUTERO_{usuario_para_nombre}_{lunes.isoformat()}.xlsx"
        return buffer, nombre_archivo
