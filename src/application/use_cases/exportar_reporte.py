from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.domain.ports.cambio_cliente_repository import CambioClienteRepository
from src.domain.ports.llamada_repository import LlamadaRepository
from src.domain.servicios.dia_visita import fecha_del_dia_en_semana

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _pintar_encabezados(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autoajustar_columnas(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


class ExportarReporte:
    def __init__(self, llamada_repo: LlamadaRepository, cambio_repo: CambioClienteRepository):
        self._llamada_repo = llamada_repo
        self._cambio_repo = cambio_repo

    async def execute(self, fecha: date, asesor: str) -> BytesIO:
        # Solo los que requieren una decisión (novedad, saltado, no contestó
        # tras 2 intentos) — los "Contestó" exitosos no aportan al reporte.
        problematicos = await self._llamada_repo.get_problematicos_del_dia(fecha, asesor)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Pendientes {fecha}"

        headers = ["Cliente", "Razón Social", "Teléfono", "Estado", "Tipo Novedad", "Observación"]
        _pintar_encabezados(ws, headers)

        for row_idx, n in enumerate(problematicos, 2):
            ws.cell(row=row_idx, column=1, value=n.get("nombre", ""))
            ws.cell(row=row_idx, column=2, value=n.get("razon_social", ""))
            ws.cell(row=row_idx, column=3, value=n.get("telefono", ""))
            ws.cell(row=row_idx, column=4, value=n.get("estado_reporte", ""))
            ws.cell(row=row_idx, column=5, value=n.get("tipo_novedad", ""))
            ws.cell(row=row_idx, column=6, value=n.get("observacion", ""))

        _autoajustar_columnas(ws)

        await self._agregar_hoja_correcciones(wb, fecha, asesor)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    async def _agregar_hoja_correcciones(self, wb, fecha: date, asesor: str) -> None:
        """
        Hoja aparte (no mezclada con novedades: son cosas distintas — las
        novedades son de un día y de clientes con incidencia; las
        correcciones pueden ser de cualquier día y de clientes que sí
        contestaron). Cubre la semana (lunes a sábado) que contiene
        `fecha`, porque el envío al jefe es semanal — misma lógica de
        semana que ya usa ExportarRuteroExcel (`fecha_del_dia_en_semana`).
        """
        lunes = fecha_del_dia_en_semana(fecha, 0)
        sabado = fecha_del_dia_en_semana(fecha, 5)

        # BR-012: solo los cambios de ESTE asesor, nunca de otro — filtrado
        # en el propio repositorio (.eq("asesor", asesor)).
        cambios = await self._cambio_repo.get_por_asesor_rango(asesor, lunes, sabado)

        ws = wb.create_sheet(title="Correcciones solicitadas")
        headers = [
            "Cod Cliente", "Cliente", "Campo", "Valor actual en ecom",
            "Valor corregido", "Asesora", "Fecha",
        ]
        _pintar_encabezados(ws, headers)

        if not cambios:
            ws.cell(
                row=2, column=1,
                value=f"Sin correcciones registradas entre {lunes.isoformat()} y {sabado.isoformat()}.",
            )
        else:
            for row_idx, c in enumerate(cambios, 2):
                cod_cliente = c.get("cod_cliente") or ""
                cell_cod = ws.cell(row=row_idx, column=1, value=str(cod_cliente) if cod_cliente != "" else "")
                cell_cod.number_format = "@"  # conserva ceros a la izquierda, igual que el rutero
                ws.cell(row=row_idx, column=2, value=c.get("nombre") or "")
                ws.cell(row=row_idx, column=3, value=c.get("campo") or "")
                ws.cell(row=row_idx, column=4, value=c.get("valor_anterior") or "")
                ws.cell(row=row_idx, column=5, value=c.get("valor_nuevo") or "")
                ws.cell(row=row_idx, column=6, value=c.get("asesor") or "")
                ws.cell(row=row_idx, column=7, value=str(c.get("fecha") or ""))

        _autoajustar_columnas(ws)
